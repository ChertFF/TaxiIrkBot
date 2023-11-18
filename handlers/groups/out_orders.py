from time import gmtime, strftime

import emoji

from data import config
from aiogram import types
from aiogram.dispatcher.filters import Command

from data.config import TIME_LIST, MORNING_TIME_LIST
from filters import IsOurChats
from utils.db_api import quick_commands
from loader import dp, bot

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet import worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from datetime import datetime



async def out_all():
    book = openpyxl.Workbook()
    sheet = book.active

    sheet['A1'] = 'ID'
    sheet['B1'] = 'Time'
    sheet['C1'] = 'Name&LastName'
    sheet['D1'] = 'City'
    sheet['E1'] = 'Address'
    sheet['F1'] = 'Number'
    sheet['G1'] = 'Ordered'
    sheet['H1'] = 'Baned'
    sheet['I1'] = 'created_at'
    sheet['J1'] = 'updated_at'
    sheet['K1'] = 'SecondAddress'
    sheet['L1'] = 'SecondOrdered'
    sheet['M1'] = 'SecondTime'
    sheet['N1'] = 'Reminder'

    row = 2
    users_ordered = await quick_commands.select_all_users()
    for user in users_ordered:
        sheet[row][0].value = user.chat_id
        sheet[row][1].value = user.time
        sheet[row][2].value = user.name
        sheet[row][3].value = user.city
        sheet[row][4].value = user.address
        sheet[row][5].value = user.number
        sheet[row][6].value = user.ordered
        sheet[row][7].value = user.ban
        sheet[row][8].value = user.created_at.strftime('%d.%m.%Y %H:%M:%S')
        sheet[row][9].value = user.updated_at.strftime('%d.%m.%Y %H:%M:%S')
        sheet[row][10].value = user.second_address
        sheet[row][11].value = user.second_time
        sheet[row][12].value = user.second_ordered
        sheet[row][13].value = user.reminder
        row += 1
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 8
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 50
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 8
    sheet.column_dimensions['H'].width = 8
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 50
    sheet.column_dimensions['L'].width = 8
    sheet.column_dimensions['M'].width = 8
    sheet.column_dimensions['N'].width = 8

    time = strftime("%m.%d.%Y_%H.%M.%S", gmtime())
    filename = f'out/out_all_{time}.xlsx'
    book.save(filename)
    book.close()

    await bot.send_document(config.ADMIN_CHAT_ID, types.InputFile(filename), caption=f'Актуальные данные для тебя {emoji.emojize(":heart_hands:")}')


@dp.message_handler(Command("out_all"), IsOurChats())
async def out_all_command(message: types.Message):
    await message.answer("Выгружаю для тебя данные по всем записям в БД...")
    await out_all()


async def out_orders(type_orders):
    book = openpyxl.Workbook()
    sheet = book.active

    sheet['A1'] = 'Время'
    sheet['B1'] = 'Город'
    sheet['C1'] = 'Адрес'
    sheet['D1'] = 'ФИО'
    sheet['E1'] = 'Номер'


    row = 2

    if type_orders == "second":
        users_ordered = await quick_commands.select_second_orders()
        try:
            users_ordered = sorted(users_ordered, key=lambda user: (MORNING_TIME_LIST.index([str(user.second_time)]), user.second_city, user.second_address))
        except:
            pass
        
        for user in users_ordered:
            sheet[row][0].value = (user.second_time)
            sheet[row][1].value = user.second_city
            sheet[row][2].value = user.second_address
            sheet[row][3].value = user.name
            sheet[row][4].value = user.number
            row += 1

    elif type_orders == "first":
        users_ordered = await quick_commands.select_orders()
        try:
            users_ordered = sorted(users_ordered, key=lambda user: (TIME_LIST.index([str(user.time)]), user.city, user.address))
        except:
            pass
        for user in users_ordered:
            sheet[row][0].value = (str(user.time))
            sheet[row][1].value = user.city
            sheet[row][2].value = user.address
            sheet[row][3].value = user.name
            sheet[row][4].value = user.number
            row += 1

    sheet.column_dimensions['A'].width = 15
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 40
    sheet.column_dimensions['D'].width = 40
    sheet.column_dimensions['E'].width = 20

    # Установка шрифта и размера шрифта
    font = Font(name='Calibri', size=12)
    sheet.font = font

    # Установка выравнивания по центру для столбцов A, B, C, D
    align_center = Alignment(horizontal='center', vertical='center')
    for column in ['A', 'B', 'C', 'D', 'E']:
        for row in range(1, sheet.max_row + 1):
            cell = sheet[column + str(row)]
            cell.alignment = align_center

    # Установка выравнивания по правому краю для столбца E
    align_right = Alignment(horizontal='right', vertical='center')
    for row in range(2, sheet.max_row + 1):
        cell = sheet['E' + str(row)]
        cell.alignment = align_right

    # Установка границ для всех ячеек, начиная с первой строки
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=column)
            cell.border = border

    # # Проходимся по значениям в столбце E, начиная с третьей строки
    # for row in sheet.iter_rows(min_row=3, min_col=5, max_col=5):
    #     for cell in row:
    #         # Получаем значение ячейки
    #         value = cell.value
    #
    #         if value is not None:
    #             # Парсим значение в формат времени
    #             time = datetime.strptime(value, '%H:%M:%S')
    #             time_serial = (time - datetime(1899, 12, 30)).total_seconds() / (24 * 60 * 60)
    #
    #             # Записываем отформатированное значение обратно в ячейку
    #             cell.value = time_serial
    #             cell.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[19]


    # # Сортировка данных по времени, населенному пункту и улице
    # data.sort_values(by=['Город', 'Адрес'], inplace=True)
    #
    # # Определение порядка сортировки по времени
    # time_order = [time[0] for time in TIME_LIST]
    #
    # # Создание временного столбца для сортировки по порядку времени
    # data["time_order"] = pd.Categorical(data['Время'], categories=time_order, ordered=True)
    #
    # # Сортировка данных по времени в заданном порядке
    # data.sort_values(by='time_order', inplace=True)
    #
    # # Удаление временного столбца
    # data.drop('time_order', axis=1, inplace=True)
    #
    # # Очистка содержимого листа
    # sheet.delete_rows(1, sheet.max_row)
    #
    # # Запись отсортированных данных обратно в лист
    # for row in dataframe_to_rows(data, index=False, header=True):
    #     sheet.append(row)

    time = strftime("%m.%d.%Y_%H.%M.%S", gmtime())
    filename = f'out/{type_orders}_orders_out_{time}.xlsx'
    book.save(filename)
    book.close()
    await bot.send_document(config.ADMIN_CHAT_ID, types.InputFile(filename), caption=f'Актуальные данные для тебя {emoji.emojize(":heart_hands:")}')


@dp.message_handler(Command("orders_out"), IsOurChats())
async def out_orders_command(message: types.Message):
    await message.answer("Выгружаю для тебя данные по вечерним заказам...")
    await out_orders("first")


@dp.message_handler(Command("morning_orders_out"), IsOurChats())
async def out_orders_command(message: types.Message):
    await message.answer("Актуальные данные для тебя по утренним заказам...")
    await out_orders("second")