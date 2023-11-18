from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from openpyxl import load_workbook

from filters import IsOurChats
from loader import dp, bot
from states import ChangeCollumn
from utils.db_api import quick_commands


@dp.message_handler(Command("mass_address_change"), IsOurChats())
async def change_user_info(message: types.Message, state: FSMContext):
    users_all = await quick_commands.select_all_users()
    wb = load_workbook('data.xlsx')
    ws = wb.active

    # Цикл для перебора строк в файле
    for row in ws.iter_rows(min_row=2, values_only=True):
        user_id = row[0]
        new_address = row[1]








